
import csv
import html
import json
import os
import time
import base64
from io import BytesIO
from collections import deque
from typing import List, Dict, Any, Optional

import cv2
import joblib
import numpy as np
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

try:
    from openpyxl import Workbook
except Exception:
    Workbook = None

# Try to use MediaPipe Tasks API (FaceLandmarker). If unavailable, fall back to
# MediaPipe Solutions FaceMesh which is more widely available in wheels.
USE_TASKS = False
FaceLandmarker = None
FaceLandmarkerOptions = None
_BaseOptions = None
Image = None
ImageFormat = None
face_detector = None

try:
    from mediapipe.tasks.python.vision import FaceLandmarker, FaceLandmarkerOptions
    from mediapipe.tasks.python.vision.face_landmarker import _BaseOptions
    try:
        from mediapipe.tasks.python.vision.core.image import Image, ImageFormat
    except Exception:
        from mediapipe.tasks.python.vision import Image, ImageFormat
    USE_TASKS = True
except Exception:
    try:
        from mediapipe import solutions as mps
        USE_TASKS = False
        mps_face_mesh = mps.face_mesh
    except Exception:
        raise ImportError("Neither MediaPipe Tasks nor Solutions FaceMesh could be imported. Please install mediapipe.")

# Load trained model if available
MODEL_FILE = os.path.join(os.path.dirname(__file__), "focus_model.pkl")
trained_clf = None
if os.path.exists(MODEL_FILE):
    try:
        trained_clf = joblib.load(MODEL_FILE)
        print(f"Loaded trained model from {MODEL_FILE}")
    except Exception as e:
        print(f"Failed to load model: {e}")

MODEL_PATH = os.path.join(os.path.dirname(__file__), "face_landmarker_v2.task")

# Download model if not
def download_model():
    if not os.path.exists(MODEL_PATH):
        import urllib.request
        urllib.request.urlretrieve("https://storage.googleapis.com/mediapipe-assets/face_landmarker_v2.task", MODEL_PATH)

if USE_TASKS:
    download_model()

# Initialize detector depending on available API
if USE_TASKS:
    options_init = FaceLandmarkerOptions(
        base_options=_BaseOptions(model_asset_path=MODEL_PATH),
        output_face_blendshapes=False,
        output_facial_transformation_matrixes=False,
        num_faces=5,
    )
    face_landmarker = FaceLandmarker.create_from_options(options_init)
else:
    face_mesh = mps_face_mesh.FaceMesh(static_image_mode=False, refine_landmarks=True, max_num_faces=5)

LEFT_EYE = [33, 160, 158, 133, 153, 144]
RIGHT_EYE = [362, 385, 387, 263, 373, 380]
LEFT_IRIS = [474, 475, 476, 477]
RIGHT_IRIS = [469, 470, 471, 472]
FACE_3D_IDX = [1, 152, 263, 33, 287, 57, 61, 291, 199]
FACE_2D_IDX = [1, 152, 263, 33, 287, 57, 61, 291, 199]

CALIBRATION_FRAMES = 100
STATIC_EAR_SLEEPY = 0.18
STATIC_IRIS_SLEEPY = 0.22

CONSEC_FRAMES_REQUIRED = 5

LOG_FILE = os.path.join(os.path.dirname(__file__), "focus_log.csv")
WEBCAM_MAX_DIM = 640
WEBCAM_STREAM_EVERY = 0.1
MAX_TRACKED_FACES = 5
TRACKS = {}
NEXT_TRACK_ID = 1
TRACK_MAX_MISSES = 8
TRACK_MATCH_DISTANCE_RATIO = 0.18
RECORDING_ACTIVE = False
RECORD_EVENTS = []
RECORD_EVENT_LIMIT = 5000
RECORD_SESSION_STARTED_AT = None
RECORD_SESSION_STOPPED_AT = None
EXPORT_DIR = os.path.join(os.path.dirname(__file__), "exports")

if not os.path.exists(LOG_FILE):
    with open(LOG_FILE, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["ts","ear","iris","yaw","pitch","roll","status","confidence"])

os.makedirs(EXPORT_DIR, exist_ok=True)

def reset_tracking_state():
    global TRACKS, NEXT_TRACK_ID
    TRACKS.clear()
    NEXT_TRACK_ID = 1

def clear_recording_state():
    global RECORDING_ACTIVE, RECORD_EVENTS, RECORD_SESSION_STARTED_AT, RECORD_SESSION_STOPPED_AT
    RECORDING_ACTIVE = False
    RECORD_EVENTS.clear()
    RECORD_SESSION_STARTED_AT = None
    RECORD_SESSION_STOPPED_AT = None

def format_timestamp(ts):
    if not ts:
        return "-"
    return time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(ts))

def build_recording_status():
    session_start = format_timestamp(RECORD_SESSION_STARTED_AT)
    if RECORDING_ACTIVE:
        mode = "Recording aktif"
        session_end = "berjalan"
    else:
        mode = "Recording berhenti"
        session_end = format_timestamp(RECORD_SESSION_STOPPED_AT)

    return {
        "mode": mode,
        "session_start": session_start,
        "session_end": session_end,
        "total_events": len(RECORD_EVENTS)
    }

def build_record_event_rows(limit=100):
    rows = []
    for event in RECORD_EVENTS[-limit:]:
        rows.append(event)
    return rows

def build_record_summary_rows():
    summary = {}
    for event in RECORD_EVENTS:
        person_id = event["id"]
        if person_id not in summary:
            summary[person_id] = {
                "id": person_id,
                "label": event["label"],
                "focused": 0,
                "not_focused": 0,
                "first_seen": event["timestamp"],
                "last_seen": event["timestamp"],
            }

        item = summary[person_id]
        if event["status"].startswith("Focused"):
            item["focused"] += 1
        else:
            item["not_focused"] += 1
        item["last_seen"] = event["timestamp"]

    rows = []
    for person_id in sorted(summary):
        rows.append(summary[person_id])
    return rows

def append_record_events(metrics):
    if not RECORDING_ACTIVE or not isinstance(metrics, dict):
        return

    timestamp = format_timestamp(time.time())
    for person in metrics.get("people", []):
        RECORD_EVENTS.append({
            "timestamp": timestamp,
            "id": person.get("id", ""),
            "label": person.get("label", ""),
            "status": person.get("status", ""),
            "confidence": float(person.get("confidence", 0.0)),
        })

    if len(RECORD_EVENTS) > RECORD_EVENT_LIMIT:
        del RECORD_EVENTS[:-RECORD_EVENT_LIMIT]

def start_recording():
    global RECORDING_ACTIVE, RECORD_SESSION_STARTED_AT, RECORD_SESSION_STOPPED_AT
    RECORDING_ACTIVE = True
    RECORD_EVENTS.clear()
    RECORD_SESSION_STARTED_AT = time.time()
    RECORD_SESSION_STOPPED_AT = None
    return {
        "status": "started",
        "recording_status": build_recording_status(),
        "events": build_record_event_rows(),
        "summary": build_record_summary_rows()
    }

def stop_recording():
    global RECORDING_ACTIVE, RECORD_SESSION_STOPPED_AT
    RECORDING_ACTIVE = False
    RECORD_SESSION_STOPPED_AT = time.time()
    return {
        "status": "stopped",
        "recording_status": build_recording_status(),
        "events": build_record_event_rows(),
        "summary": build_record_summary_rows()
    }

def clear_recording():
    clear_recording_state()
    return {
        "status": "cleared",
        "recording_status": build_recording_status(),
        "events": build_record_event_rows(),
        "summary": build_record_summary_rows()
    }

def build_excel_xml_worksheet(sheet_name, headers, rows):
    lines = [f'<Worksheet ss:Name="{html.escape(sheet_name)}"><Table>']
    all_rows = [headers] + rows
    for row in all_rows:
        lines.append("<Row>")
        for cell in row:
            cell_value = "" if cell is None else str(cell)
            lines.append(
                "<Cell><Data ss:Type=\"String\">"
                f"{html.escape(cell_value)}"
                "</Data></Cell>"
            )
        lines.append("</Row>")
    lines.append("</Table></Worksheet>")
    return "".join(lines)

def export_record_to_excel_xml(export_path):
    event_rows = build_record_event_rows(limit=RECORD_EVENT_LIMIT)
    summary_rows = build_record_summary_rows()
    session_rows = [
        ["recording_status", "aktif" if RECORDING_ACTIVE else "berhenti"],
        ["session_started_at", format_timestamp(RECORD_SESSION_STARTED_AT)],
        ["session_stopped_at", format_timestamp(RECORD_SESSION_STOPPED_AT)],
        ["total_events", len(RECORD_EVENTS)],
    ]

    workbook_xml = [
        "<?xml version=\"1.0\"?>",
        "<?mso-application progid=\"Excel.Sheet\"?>",
        "<Workbook xmlns=\"urn:schemas-microsoft-com:office:spreadsheet\"",
        " xmlns:o=\"urn:schemas-microsoft-com:office:office\"",
        " xmlns:x=\"urn:schemas-microsoft-com:office:excel\"",
        " xmlns:ss=\"urn:schemas-microsoft-com:office:spreadsheet\">",
        build_excel_xml_worksheet("session_info", ["field", "value"], session_rows),
        build_excel_xml_worksheet("events", ["timestamp", "id", "label", "status", "confidence"], 
                                   [[e["timestamp"], e["id"], e["label"], e["status"], e["confidence"]] for e in event_rows]),
        build_excel_xml_worksheet("summary", ["id", "label", "focused", "not_focused", "total", "first_seen", "last_seen"],
                                   [[s["id"], s["label"], s["focused"], s["not_focused"], 
                                     s["focused"] + s["not_focused"], s["first_seen"], s["last_seen"]] for s in summary_rows]),
        "</Workbook>",
    ]

    with open(export_path, "w", encoding="utf-8") as export_file:
        export_file.write("".join(workbook_xml))

def export_record_to_excel():
    timestamp = time.strftime("%Y%m%d_%H%M%S", time.localtime(time.time()))
    if Workbook is not None:
        export_path = os.path.abspath(os.path.join(EXPORT_DIR, f"focus_record_{timestamp}.xlsx"))
        workbook = Workbook()
        session_sheet = workbook.active
        session_sheet.title = "session_info"
        session_sheet.append(["field", "value"])
        session_sheet.append(["recording_status", "aktif" if RECORDING_ACTIVE else "berhenti"])
        session_sheet.append(["session_started_at", format_timestamp(RECORD_SESSION_STARTED_AT)])
        session_sheet.append(["session_stopped_at", format_timestamp(RECORD_SESSION_STOPPED_AT)])
        session_sheet.append(["total_events", len(RECORD_EVENTS)])

        events_sheet = workbook.create_sheet(title="events")
        events_sheet.append(["timestamp", "id", "label", "status", "confidence"])
        for e in build_record_event_rows(limit=RECORD_EVENT_LIMIT):
            events_sheet.append([e["timestamp"], e["id"], e["label"], e["status"], e["confidence"]])

        summary_sheet = workbook.create_sheet(title="summary")
        summary_sheet.append(["id", "label", "focused", "not_focused", "total", "first_seen", "last_seen"])
        for s in build_record_summary_rows():
            summary_sheet.append([s["id"], s["label"], s["focused"], s["not_focused"],
                                  s["focused"] + s["not_focused"], s["first_seen"], s["last_seen"]])

        workbook.save(export_path)
        return {"status": "success", "filename": os.path.basename(export_path), "path": export_path}

    export_path = os.path.abspath(os.path.join(EXPORT_DIR, f"focus_record_{timestamp}.xls"))
    export_record_to_excel_xml(export_path)
    return {"status": "success", "filename": os.path.basename(export_path), "path": export_path}

def eye_aspect_ratio(landmarks, eye_indices):
    points = [(landmarks[i].x, landmarks[i].y) for i in eye_indices]
    vertical1 = np.linalg.norm(np.array(points[1]) - np.array(points[5]))
    vertical2 = np.linalg.norm(np.array(points[2]) - np.array(points[4]))
    horizontal = np.linalg.norm(np.array(points[0]) - np.array(points[3]))
    ear = (vertical1 + vertical2) / (2.0 * horizontal)
    return ear

def iris_aspect_ratio(landmarks, iris_indices):
    xs = [landmarks[i].x for i in iris_indices]
    ys = [landmarks[i].y for i in iris_indices]
    width = max(xs) - min(xs)
    height = max(ys) - min(ys)
    return height / width if width > 0 else 0

def head_pose(landmarks, w, h):
    model_points = np.array([
        [0.0, 0.0, 0.0],
        [0.0, -63.6, -12.5],
        [-43.3, 32.7, -26.0],
        [43.3, 32.7, -26.0],
        [-28.9, -28.9, -24.1],
        [28.9, -28.9, -24.1],
        [-61.6, -11.2, -39.5],
        [61.6, -11.2, -39.5],
        [0.0, -48.0, -50.0],
    ])
    image_points = np.array([
        [landmarks[i].x * w, landmarks[i].y * h] for i in FACE_2D_IDX
    ], dtype='double')
    focal_length = w
    center = (w / 2, h / 2)
    camera_matrix = np.array(
        [[focal_length, 0, center[0]],
         [0, focal_length, center[1]],
         [0, 0, 1]], dtype='double')
    dist_coeffs = np.zeros((4, 1))
    success, rotation_vector, translation_vector = cv2.solvePnP(
        model_points, image_points, camera_matrix, dist_coeffs, flags=cv2.SOLVEPNP_ITERATIVE)
    if not success:
        return 0, 0, 0
    rmat, _ = cv2.Rodrigues(rotation_vector)
    sy = np.sqrt(rmat[0, 0] ** 2 + rmat[1, 0] ** 2)
    x = np.arctan2(rmat[2, 1], rmat[2, 2])
    y = np.arctan2(-rmat[2, 0], sy)
    z = np.arctan2(rmat[1, 0], rmat[0, 0])
    return np.degrees(x), np.degrees(y), np.degrees(z)

def create_track_state():
    return {
        "ear_history": deque(maxlen=10),
        "iris_history": deque(maxlen=10),
        "pose_history": deque(maxlen=10),
        "calibration_mode": False,
        "calib_ear": [],
        "calib_iris": [],
        "ear_thresh": 0.20,
        "iris_thresh": 0.25,
        "closed_count": 0,
        "open_count": 0,
        "centroid": None,
        "misses": 0,
        "last_seen": 0.0,
    }

def get_face_bbox(landmarks, width, height):
    xs = [landmark.x for landmark in landmarks]
    ys = [landmark.y for landmark in landmarks]
    x1 = max(0, int(min(xs) * width))
    y1 = max(0, int(min(ys) * height))
    x2 = min(width - 1, int(max(xs) * width))
    y2 = min(height - 1, int(max(ys) * height))
    return x1, y1, x2, y2

def get_face_landmarks_list(rgb_frame):
    if USE_TASKS:
        mp_image = Image(image_format=ImageFormat.SRGB, data=rgb_frame)
        result = face_landmarker.detect(mp_image)
        return list(result.face_landmarks or [])

    results = face_mesh.process(rgb_frame)
    if results and results.multi_face_landmarks:
        return [face_landmarks.landmark for face_landmarks in results.multi_face_landmarks]
    return []

def assign_track_id(centroid, frame_shape, used_track_ids):
    global NEXT_TRACK_ID

    frame_height, frame_width = frame_shape[:2]
    max_distance = max(frame_width, frame_height) * TRACK_MATCH_DISTANCE_RATIO
    best_track_id = None
    best_distance = max_distance

    for track_id, track_state in TRACKS.items():
        if track_id in used_track_ids:
            continue

        previous_centroid = track_state.get("centroid")
        if previous_centroid is None:
            continue

        distance = np.linalg.norm(np.array(centroid) - np.array(previous_centroid))
        if distance < best_distance:
            best_distance = distance
            best_track_id = track_id

    if best_track_id is None and len(TRACKS) < MAX_TRACKED_FACES:
        best_track_id = NEXT_TRACK_ID
        NEXT_TRACK_ID += 1
        TRACKS[best_track_id] = create_track_state()
    elif best_track_id is None:
        stale_track_id = max(TRACKS, key=lambda track_id: TRACKS[track_id]["misses"])
        best_track_id = stale_track_id
        TRACKS[best_track_id] = create_track_state()

    track_state = TRACKS[best_track_id]
    track_state["centroid"] = centroid
    track_state["misses"] = 0
    track_state["last_seen"] = time.time()
    used_track_ids.add(best_track_id)

    return best_track_id, track_state

def cleanup_tracks():
    stale_track_ids = [
        track_id for track_id, track_state in TRACKS.items()
        if track_state["misses"] > TRACK_MAX_MISSES
    ]
    for track_id in stale_track_ids:
        TRACKS.pop(track_id, None)

def analyze_face(landmarks, frame_shape, track_id, track_state, use_trained=True):
    frame_height, frame_width = frame_shape[:2]
    left_ear = eye_aspect_ratio(landmarks, LEFT_EYE)
    right_ear = eye_aspect_ratio(landmarks, RIGHT_EYE)
    left_iris_ar = iris_aspect_ratio(landmarks, LEFT_IRIS)
    right_iris_ar = iris_aspect_ratio(landmarks, RIGHT_IRIS)
    ear = (left_ear + right_ear) / 2
    iris_ar = (left_iris_ar + right_iris_ar) / 2
    yaw, pitch, roll = head_pose(landmarks, frame_width, frame_height)

    track_state["ear_history"].append(ear)
    track_state["iris_history"].append(iris_ar)
    track_state["pose_history"].append((yaw, pitch, roll))
    ear_smooth = np.mean(track_state["ear_history"])
    iris_smooth = np.mean(track_state["iris_history"])
    pose_smooth = np.mean(track_state["pose_history"], axis=0)

    if track_state["calibration_mode"] or (
        len(track_state["calib_ear"]) < CALIBRATION_FRAMES and ear_smooth > 0.25
    ):
        track_state["calibration_mode"] = True
        track_state["calib_ear"].append(ear_smooth)
        track_state["calib_iris"].append(iris_smooth)
        if len(track_state["calib_ear"]) >= CALIBRATION_FRAMES:
            track_state["ear_thresh"] = np.mean(track_state["calib_ear"]) * 0.8
            track_state["iris_thresh"] = np.mean(track_state["calib_iris"]) * 0.8
            track_state["calibration_mode"] = False
            track_state["calib_ear"].clear()
            track_state["calib_iris"].clear()

    history_len = len(track_state["ear_history"])
    both_eyes_closed = (
        ear_smooth < track_state["ear_thresh"] and iris_smooth < track_state["iris_thresh"]
    )
    turned_away = abs(pose_smooth[1]) > 30
    if both_eyes_closed:
        track_state["closed_count"] += 1
        track_state["open_count"] = 0
    else:
        track_state["open_count"] += 1
        track_state["closed_count"] = 0

    if history_len <= 2:
        is_sleepy = (ear_smooth < STATIC_EAR_SLEEPY and iris_smooth < STATIC_IRIS_SLEEPY)
    else:
        is_sleepy = track_state["closed_count"] >= CONSEC_FRAMES_REQUIRED

    left_iris_x = np.mean([landmarks[i].x for i in LEFT_IRIS])
    right_iris_x = np.mean([landmarks[i].x for i in RIGHT_IRIS])
    left_eye_x_center = np.mean([landmarks[i].x for i in LEFT_EYE])
    right_eye_x_center = np.mean([landmarks[i].x for i in RIGHT_EYE])
    left_eye_w = max([landmarks[i].x for i in LEFT_EYE]) - min([landmarks[i].x for i in LEFT_EYE])
    right_eye_w = max([landmarks[i].x for i in RIGHT_EYE]) - min([landmarks[i].x for i in RIGHT_EYE])
    if left_eye_w < 1e-6:
        left_eye_w = 1e-6
    if right_eye_w < 1e-6:
        right_eye_w = 1e-6
    left_offset = (left_iris_x - left_eye_x_center) / left_eye_w
    right_offset = (right_iris_x - right_eye_x_center) / right_eye_w
    gaze_x = (left_offset + right_offset) / 2.0

    if use_trained and trained_clf is not None:
        try:
            features = np.array(
                [ear_smooth, iris_smooth, pose_smooth[0], pose_smooth[1], pose_smooth[2], gaze_x]
            )
            pred = trained_clf.predict([features])[0]
            pred_proba = trained_clf.predict_proba([features])[0]
            if pred == 1:
                status = "Focused"
                confidence = float(pred_proba[1])
            else:
                status = "Not Focused"
                confidence = float(pred_proba[0])
        except Exception as e:
            print(f"Model prediction error: {e}")
            if is_sleepy:
                status = "Not Focused (Sleepy)"
                confidence = 0.5
            elif turned_away:
                status = "Not Focused (Turned Away)"
                confidence = 0.6
            else:
                status = "Focused"
                confidence = 1.0
    else:
        if is_sleepy:
            status = "Not Focused (Sleepy)"
            confidence = 0.5
        elif turned_away:
            status = "Not Focused (Turned Away)"
            confidence = 0.6
        else:
            status = "Focused"
            confidence = 1.0

    bbox = get_face_bbox(landmarks, frame_width, frame_height)
    metrics = {
        "id": int(track_id),
        "label": f"Person {track_id}",
        "bbox": {
            "x1": int(bbox[0]),
            "y1": int(bbox[1]),
            "x2": int(bbox[2]),
            "y2": int(bbox[3]),
        },
        "ear": float(ear_smooth),
        "iris": float(iris_smooth),
        "yaw": float(pose_smooth[0]),
        "pitch": float(pose_smooth[1]),
        "roll": float(pose_smooth[2]),
        "gaze_x": float(gaze_x),
        "ear_thresh": float(track_state["ear_thresh"]),
        "iris_thresh": float(track_state["iris_thresh"]),
        "status": status,
        "confidence": confidence,
    }

    try:
        with open(LOG_FILE, "a", newline="") as f:
            writer = csv.writer(f)
            writer.writerow([
                time.time(),
                ear_smooth,
                iris_smooth,
                pose_smooth[0],
                pose_smooth[1],
                pose_smooth[2],
                f"P{track_id}:{status}",
                confidence,
            ])
    except Exception:
        pass

    return metrics

def draw_person_overlay(frame_bgr, person_metrics):
    bbox = person_metrics["bbox"]
    status = person_metrics["status"]
    color = (0, 255, 0) if status.startswith("Focused") else (0, 0, 255)

    x1, y1, x2, y2 = bbox["x1"], bbox["y1"], bbox["x2"], bbox["y2"]
    cv2.rectangle(frame_bgr, (x1, y1), (x2, y2), color, 2)

    label = (
        f"{person_metrics['label']} | {status} | "
        f"{person_metrics['confidence']:.2f}"
    )
    text_y = max(25, y1 - 10)
    cv2.putText(
        frame_bgr,
        label,
        (x1, text_y),
        cv2.FONT_HERSHEY_SIMPLEX,
        0.6,
        color,
        2,
    )

def detect_focus(rgb_frame, use_trained=True, record=False):
    for track_state in TRACKS.values():
        track_state["misses"] += 1

    landmarks_list = get_face_landmarks_list(rgb_frame)
    bgr = cv2.cvtColor(rgb_frame, cv2.COLOR_RGB2BGR)
    frame_height, frame_width = bgr.shape[:2]

    if not landmarks_list:
        cleanup_tracks()
        summary = {
            "status": "no_face",
            "people_count": 0,
            "focused_count": 0,
            "not_focused_count": 0,
            "people": []
        }
        if record:
            append_record_events(summary)
        return summary

    used_track_ids = set()
    people_metrics = []
    sorted_landmarks = sorted(
        landmarks_list,
        key=lambda landmarks: get_face_bbox(landmarks, frame_width, frame_height)[0],
    )

    for landmarks in sorted_landmarks:
        bbox = get_face_bbox(landmarks, frame_width, frame_height)
        centroid = ((bbox[0] + bbox[2]) / 2.0, (bbox[1] + bbox[3]) / 2.0)
        track_id, track_state = assign_track_id(centroid, bgr.shape, used_track_ids)
        person_metrics = analyze_face(
            landmarks,
            bgr.shape,
            track_id,
            track_state,
            use_trained=use_trained,
        )
        people_metrics.append(person_metrics)

    cleanup_tracks()

    focused_count = sum(1 for person in people_metrics if person["status"].startswith("Focused"))
    summary = {
        "status": "multi_face",
        "people_count": len(people_metrics),
        "focused_count": focused_count,
        "not_focused_count": len(people_metrics) - focused_count,
        "people": people_metrics
    }

    if record:
        append_record_events(summary)

    return summary


# Pydantic models
class AnalyzeFrameRequest(BaseModel):
    image_base64: str
    use_trained: bool = True
    record: bool = False


app = FastAPI(title="Focus Detector API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/health")
async def health_check():
    return {"status": "ok"}


@app.post("/focus/analyze-frame")
async def analyze_frame(request: AnalyzeFrameRequest):
    try:
        # Decode base64
        image_data = base64.b64decode(request.image_base64.split(",")[1] if "," in request.image_base64 else request.image_base64)
        nparr = np.frombuffer(image_data, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        if img is None:
            raise HTTPException(status_code=400, detail="Could not decode image")
        rgb_frame = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

        summary = detect_focus(rgb_frame, use_trained=request.use_trained, record=request.record)
        return summary
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/focus/record/start")
async def start_record():
    return start_recording()


@app.post("/focus/record/stop")
async def stop_record():
    return stop_recording()


@app.post("/focus/record/clear")
async def clear_record():
    return clear_recording()


@app.get("/focus/record/status")
async def get_record_status():
    return {
        "recording_status": build_recording_status(),
        "events": build_record_event_rows(),
        "summary": build_record_summary_rows()
    }


@app.get("/focus/record/export")
async def export_record():
    result = export_record_to_excel()
    if not os.path.exists(result["path"]):
        raise HTTPException(status_code=404, detail="Export file not found")
    
    from fastapi.responses import FileResponse
    return FileResponse(
        path=result["path"],
        filename=result["filename"],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.post("/focus/reset")
async def reset_state():
    reset_tracking_state()
    return {"status": "reset"}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)

