import csv
import json
import os
import time
import base64
import urllib.request
from collections import deque
from types import SimpleNamespace
import cv2
import gradio as gr
import joblib
import numpy as np
from fastapi import FastAPI, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Dict, List, Optional, Any
try:
    from openpyxl import Workbook
except Exception:
    Workbook = None


# Initialize FastAPI app
app = FastAPI(title="Focus Detection API")

DEBUG_ENV_PATH = r"c:\Users\LENOVO\Documents\fokus_monitoring_yolo_MERN\.dbg\face-distance-detection.env"


def report_debug_event(hypothesis_id, location, msg, data=None, run_id="pre-fix"):
    # #region debug-point A:report-debug-event
    _u, _s = "http://127.0.0.1:7777/event", "face-distance-detection"
    try:
        with open(DEBUG_ENV_PATH, "r", encoding="utf-8") as _f:
            for _line in _f:
                if _line.startswith("DEBUG_SERVER_URL="):
                    _u = _line.split("=", 1)[1].strip()
                elif _line.startswith("DEBUG_SESSION_ID="):
                    _s = _line.split("=", 1)[1].strip()
    except Exception:
        pass
    try:
        urllib.request.urlopen(
            urllib.request.Request(
                _u,
                data=json.dumps(
                    {
                        "sessionId": _s,
                        "runId": run_id,
                        "hypothesisId": hypothesis_id,
                        "location": location,
                        "msg": msg,
                        "data": data or {},
                        "ts": int(time.time() * 1000),
                    }
                ).encode(),
                headers={"Content-Type": "application/json"},
            ),
            timeout=0.6,
        ).read()
    except Exception:
        pass
    # #endregion

# Pydantic models for API requests
class FrameRequest(BaseModel):
    image_base64: str
    use_trained: bool = True


# Try to use MediaPipe Tasks API (FaceLandmarker). If unavailable, fall back to
# MediaPipe Solutions FaceMesh which is more widely available in wheels.
USE_TASKS = False
FaceLandmarker = None
FaceLandmarkerOptions = None
_BaseOptions = None
Image = None
ImageFormat = None
face_detector = None
mp_face_detection = None

try:
    from mediapipe.tasks.python.vision import FaceLandmarker, FaceLandmarkerOptions
    from mediapipe.tasks.python.vision.face_landmarker import _BaseOptions
    try:
        from mediapipe.tasks.python.vision.core.image import Image, ImageFormat
    except Exception:
        from mediapipe.tasks.python.vision import Image, ImageFormat
    from mediapipe import solutions as mps
    mp_face_detection = mps.face_detection
    USE_TASKS = True
except Exception:
    # fallback to solutions FaceMesh
    try:
        from mediapipe import solutions as mps
        USE_TASKS = False
        mps_face_mesh = mps.face_mesh
        mp_face_detection = mps.face_detection
    except Exception:
        raise ImportError("Neither MediaPipe Tasks nor Solutions FaceMesh could be imported. Please install mediapipe.")

# Load trained model if available
MODEL_FILE = "focus_model.pkl"
trained_clf = None
if os.path.exists(MODEL_FILE):
    try:
        trained_clf = joblib.load(MODEL_FILE)
        print(f"Loaded trained model from {MODEL_FILE}")
    except Exception as e:
        print(f"Failed to load model: {e}")

MODEL_PATH = "face_landmarker_v2.task"
MODEL_URL = "https://storage.googleapis.com/mediapipe-assets/face_landmarker_v2.task"

# Download model if not
def download_model():
    if not os.path.exists(MODEL_PATH):
        import urllib.request
        urllib.request.urlretrieve(MODEL_URL, MODEL_PATH)

if USE_TASKS:
    download_model()

FACE_DETECTION_MIN_CONFIDENCE = 0.35

# Initialize detector depending on available API
if USE_TASKS:
    options_init = FaceLandmarkerOptions(
        base_options=_BaseOptions(model_asset_path=MODEL_PATH),
        output_face_blendshapes=False,
        output_facial_transformation_matrixes=False,
        num_faces=5,
    )
    face_landmarker = FaceLandmarker.create_from_options(options_init)
else:
    # use Solutions FaceMesh as fallback
    face_mesh = mps_face_mesh.FaceMesh(static_image_mode=False, refine_landmarks=True, max_num_faces=5)

if mp_face_detection is not None:
    face_detector = mp_face_detection.FaceDetection(
        model_selection=1,
        min_detection_confidence=FACE_DETECTION_MIN_CONFIDENCE,
    )

LEFT_EYE = [33, 160, 158, 133, 153, 144]
RIGHT_EYE = [362, 385, 387, 263, 373, 380]
LEFT_IRIS = [474, 475, 476, 477]
RIGHT_IRIS = [469, 470, 471, 472]
FACE_3D_IDX = [1, 152, 263, 33, 287, 57, 61, 291, 199]
FACE_2D_IDX = [1, 152, 263, 33, 287, 57, 61, 291, 199]

CALIBRATION_FRAMES = 100
STATIC_EAR_SLEEPY = 0.18
STATIC_IRIS_SLEEPY = 0.22

CONSEC_FRAMES_REQUIRED = 5

LOG_FILE = "focus_log.csv"
WEBCAM_MAX_DIM = 1280
WEBCAM_STREAM_EVERY = 0.1
MAX_TRACKED_FACES = 5
TRACKS = {}
NEXT_TRACK_ID = 1
TRACK_MAX_MISSES = 18
TRACK_MATCH_DISTANCE_RATIO = 0.28
TRACK_MIN_IOU = 0.08
DETECTION_RETRY_MAX_DIM = 1280
FACE_DETECTION_PADDING_RATIO = 0.22
RECORDING_ACTIVE = False
RECORD_EVENTS = []
RECORD_EVENT_LIMIT = 5000
RECORD_SESSION_STARTED_AT = None
RECORD_SESSION_STOPPED_AT = None
EXPORT_DIR = "exports"
if not os.path.exists(LOG_FILE):
    with open(LOG_FILE, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["ts","ear","iris","yaw","pitch","roll","status","confidence"])

def reset_tracking_state():
    global TRACKS, NEXT_TRACK_ID

    TRACKS.clear()
    NEXT_TRACK_ID = 1


def clear_recording_state():
    global RECORDING_ACTIVE, RECORD_EVENTS, RECORD_SESSION_STARTED_AT, RECORD_SESSION_STOPPED_AT

    RECORDING_ACTIVE = False
    RECORD_EVENTS.clear()
    RECORD_SESSION_STARTED_AT = None
    RECORD_SESSION_STOPPED_AT = None


def format_timestamp(ts):
    if not ts:
        return "-"
    return time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(ts))


def build_recording_status():
    session_start = format_timestamp(RECORD_SESSION_STARTED_AT)
    if RECORDING_ACTIVE:
        mode = "Recording aktif"
        session_end = "berjalan"
    else:
        mode = "Recording berhenti"
        session_end = format_timestamp(RECORD_SESSION_STOPPED_AT)

    return (
        f"{mode} | Mulai: {session_start} | Selesai: {session_end} | "
        f"Total event: {len(RECORD_EVENTS)}"
    )


def build_record_event_rows(limit=100):
    rows = []
    for event in RECORD_EVENTS[-limit:]:
        rows.append(
            [
                event["timestamp"],
                event["id"],
                event["label"],
                event["status"],
                round(float(event["confidence"]), 3),
            ]
        )
    return rows


def build_record_summary_rows():
    summary = {}
    for event in RECORD_EVENTS:
        person_id = event["id"]
        if person_id not in summary:
            summary[person_id] = {
                "id": person_id,
                "label": event["label"],
                "focused": 0,
                "not_focused": 0,
                "first_seen": event["timestamp"],
                "last_seen": event["timestamp"],
            }

        item = summary[person_id]
        if event["status"].startswith("Focused"):
            item["focused"] += 1
        else:
            item["not_focused"] += 1
        item["last_seen"] = event["timestamp"]

    rows = []
    for person_id in sorted(summary):
        item = summary[person_id]
        total = item["focused"] + item["not_focused"]
        rows.append(
            [
                item["id"],
                item["label"],
                item["focused"],
                item["not_focused"],
                total,
                item["first_seen"],
                item["last_seen"],
            ]
        )
    return rows


def append_record_events(metrics):
    if not RECORDING_ACTIVE or not isinstance(metrics, dict):
        return

    timestamp = format_timestamp(time.time())
    for person in metrics.get("people", []):
        RECORD_EVENTS.append(
            {
                "timestamp": timestamp,
                "id": person.get("id", ""),
                "label": person.get("label", ""),
                "status": person.get("status", ""),
                "confidence": float(person.get("confidence", 0.0)),
            }
        )

    if len(RECORD_EVENTS) > RECORD_EVENT_LIMIT:
        del RECORD_EVENTS[:-RECORD_EVENT_LIMIT]


def start_recording():
    global RECORDING_ACTIVE, RECORD_SESSION_STARTED_AT, RECORD_SESSION_STOPPED_AT

    reset_tracking_state()
    RECORDING_ACTIVE = True
    RECORD_EVENTS.clear()
    RECORD_SESSION_STARTED_AT = time.time()
    RECORD_SESSION_STOPPED_AT = None
    return build_recording_status(), build_record_event_rows(), build_record_summary_rows()


def stop_recording():
    global RECORDING_ACTIVE, RECORD_SESSION_STOPPED_AT

    RECORDING_ACTIVE = False
    RECORD_SESSION_STOPPED_AT = time.time()
    return build_recording_status(), build_record_event_rows(), build_record_summary_rows()


def clear_recording():
    clear_recording_state()
    return build_recording_status(), build_record_event_rows(), build_record_summary_rows()


def export_record_to_excel():
    if Workbook is None:
        raise RuntimeError("openpyxl tidak tersedia. Export Excel .xlsx tidak dapat dijalankan.")

    os.makedirs(EXPORT_DIR, exist_ok=True)
    timestamp = time.strftime("%Y%m%d_%H%M%S", time.localtime(time.time()))
    export_path = os.path.abspath(os.path.join(EXPORT_DIR, f"focus_record_{timestamp}.xlsx"))
    workbook = Workbook()
    session_sheet = workbook.active
    session_sheet.title = "session_info"
    session_sheet.append(["field", "value"])
    session_sheet.append(["recording_status", "aktif" if RECORDING_ACTIVE else "berhenti"])
    session_sheet.append(["session_started_at", format_timestamp(RECORD_SESSION_STARTED_AT)])
    session_sheet.append(["session_stopped_at", format_timestamp(RECORD_SESSION_STOPPED_AT)])
    session_sheet.append(["total_events", len(RECORD_EVENTS)])

    events_sheet = workbook.create_sheet(title="events")
    events_sheet.append(["timestamp", "id", "label", "status", "confidence"])
    for row in build_record_event_rows(limit=RECORD_EVENT_LIMIT):
        events_sheet.append(row)

    summary_sheet = workbook.create_sheet(title="summary")
    summary_sheet.append(["id", "label", "focused", "not_focused", "total", "first_seen", "last_seen"])
    for row in build_record_summary_rows():
        summary_sheet.append(row)

    workbook.save(export_path)
    return f"Export berhasil: {os.path.basename(export_path)}", export_path


def reset_runtime_state():
    reset_tracking_state()

    return (
        "State di-reset. Silakan arahkan wajah ke kamera untuk mulai tes lagi.",
        json.dumps(
            {
                "model_loaded": trained_clf is not None,
                "tracked_faces": 0,
            },
            indent=2,
        ),
        [],
    )


def reset_webcam_state():
    reset_tracking_state()
    clear_recording_state()
    return (
        "State webcam dan record di-reset. Silakan mulai tes lagi.",
        json.dumps(
            {
                "model_loaded": trained_clf is not None,
                "tracked_faces": 0,
                "recording": False,
            },
            indent=2,
        ),
        [],
        build_recording_status(),
        build_record_event_rows(),
        build_record_summary_rows(),
    )


def format_metrics(metrics):
    if not metrics:
        return "{}"
    return json.dumps(metrics, indent=2)


def build_people_rows(metrics):
    people = metrics.get("people", []) if isinstance(metrics, dict) else []
    rows = []
    for person in people:
        bbox = person.get("bbox", {})
        rows.append(
            [
                person.get("id", ""),
                person.get("label", ""),
                person.get("status", ""),
                round(float(person.get("confidence", 0.0)), 3),
                f"({bbox.get('x1', 0)}, {bbox.get('y1', 0)}) - ({bbox.get('x2', 0)}, {bbox.get('y2', 0)})",
                round(float(person.get("gaze_x", 0.0)), 3),
            ]
        )
    return rows


def optimize_frame_for_webcam(frame, max_dim=WEBCAM_MAX_DIM):
    if frame is None or not isinstance(frame, np.ndarray) or frame.size == 0:
        return frame

    height, width = frame.shape[:2]
    largest_dim = max(height, width)
    if largest_dim <= max_dim:
        return frame

    scale = max_dim / float(largest_dim)
    resized = cv2.resize(
        frame,
        (int(width * scale), int(height * scale)),
        interpolation=cv2.INTER_AREA,
    )
    return resized

def eye_aspect_ratio(landmarks, eye_indices):
    points = [(landmarks[i].x, landmarks[i].y) for i in eye_indices]
    vertical1 = np.linalg.norm(np.array(points[1]) - np.array(points[5]))
    vertical2 = np.linalg.norm(np.array(points[2]) - np.array(points[4]))
    horizontal = np.linalg.norm(np.array(points[0]) - np.array(points[3]))
    ear = (vertical1 + vertical2) / (2.0 * horizontal)
    return ear

def iris_aspect_ratio(landmarks, iris_indices):
    xs = [landmarks[i].x for i in iris_indices]
    ys = [landmarks[i].y for i in iris_indices]
    width = max(xs) - min(xs)
    height = max(ys) - min(ys)
    return height / width if width > 0 else 0


def head_pose(landmarks, w, h):
    model_points = np.array([
        [0.0, 0.0, 0.0],
        [0.0, -63.6, -12.5],
        [-43.3, 32.7, -26.0],
        [43.3, 32.7, -26.0],
        [-28.9, -28.9, -24.1],
        [28.9, -28.9, -24.1],
        [-61.6, -11.2, -39.5],
        [61.6, -11.2, -39.5],
        [0.0, -48.0, -50.0],
    ])
    image_points = np.array([
        [landmarks[i].x * w, landmarks[i].y * h] for i in FACE_2D_IDX
    ], dtype='double')
    focal_length = w
    center = (w / 2, h / 2)
    camera_matrix = np.array(
        [[focal_length, 0, center[0]],
         [0, focal_length, center[1]],
         [0, 0, 1]], dtype='double')
    dist_coeffs = np.zeros((4, 1))
    success, rotation_vector, translation_vector = cv2.solvePnP(
        model_points, image_points, camera_matrix, dist_coeffs, flags=cv2.SOLVEPNP_ITERATIVE)
    if not success:
        return 0, 0, 0
    rmat, _ = cv2.Rodrigues(rotation_vector)
    sy = np.sqrt(rmat[0, 0] ** 2 + rmat[1, 0] ** 2)
    x = np.arctan2(rmat[2, 1], rmat[2, 2])
    y = np.arctan2(-rmat[2, 0], sy)
    z = np.arctan2(rmat[1, 0], rmat[0, 0])
    return np.degrees(x), np.degrees(y), np.degrees(z)

def create_track_state():
    return {
        "ear_history": deque(maxlen=10),
        "iris_history": deque(maxlen=10),
        "pose_history": deque(maxlen=10),
        "calibration_mode": False,
        "calib_ear": [],
        "calib_iris": [],
        "ear_thresh": 0.20,
        "iris_thresh": 0.25,
        "closed_count": 0,
        "open_count": 0,
        "centroid": None,
        "bbox": None,
        "misses": 0,
        "last_seen": 0.0,
    }


def get_face_bbox(landmarks, width, height):
    xs = [landmark.x for landmark in landmarks]
    ys = [landmark.y for landmark in landmarks]

    x1 = max(0, int(min(xs) * width))
    y1 = max(0, int(min(ys) * height))
    x2 = min(width - 1, int(max(xs) * width))
    y2 = min(height - 1, int(max(ys) * height))

    return x1, y1, x2, y2


def get_face_landmarks_list(rgb_frame):
    if USE_TASKS:
        mp_image = Image(image_format=ImageFormat.SRGB, data=rgb_frame)
        result = face_landmarker.detect(mp_image)
        return list(result.face_landmarks or [])

    results = face_mesh.process(rgb_frame)
    if results and results.multi_face_landmarks:
        return [face_landmarks.landmark for face_landmarks in results.multi_face_landmarks]
    return []


def detect_face_regions(rgb_frame):
    if face_detector is None:
        return []

    results = face_detector.process(rgb_frame)
    detections = getattr(results, "detections", None) or []
    frame_height, frame_width = rgb_frame.shape[:2]
    face_regions = []

    for detection in detections:
        relative_bbox = detection.location_data.relative_bounding_box
        x1 = max(0, int(relative_bbox.xmin * frame_width))
        y1 = max(0, int(relative_bbox.ymin * frame_height))
        bbox_w = max(1, int(relative_bbox.width * frame_width))
        bbox_h = max(1, int(relative_bbox.height * frame_height))
        x2 = min(frame_width - 1, x1 + bbox_w)
        y2 = min(frame_height - 1, y1 + bbox_h)
        score = float(detection.score[0]) if getattr(detection, "score", None) else 0.0
        face_regions.append(
            {
                "bbox": (x1, y1, x2, y2),
                "score": score,
            }
        )

    return sorted(face_regions, key=lambda region: region["bbox"][0])


def upscale_frame_for_detection(rgb_frame, target_max_dim=DETECTION_RETRY_MAX_DIM):
    frame_height, frame_width = rgb_frame.shape[:2]
    largest_dim = max(frame_width, frame_height)
    if largest_dim >= target_max_dim:
        return rgb_frame, 1.0

    scale = target_max_dim / float(largest_dim)
    resized = cv2.resize(
        rgb_frame,
        (max(1, int(frame_width * scale)), max(1, int(frame_height * scale))),
        interpolation=cv2.INTER_CUBIC,
    )
    return resized, scale


def detect_landmarks_direct_with_retry(rgb_frame):
    landmarks_list = get_face_landmarks_list(rgb_frame)
    if landmarks_list:
        return landmarks_list, rgb_frame, 1.0

    retry_frame, retry_scale = upscale_frame_for_detection(rgb_frame)
    if retry_scale <= 1.01:
        return [], rgb_frame, 1.0

    retry_landmarks_list = get_face_landmarks_list(retry_frame)
    if retry_landmarks_list:
        return retry_landmarks_list, retry_frame, retry_scale

    return [], rgb_frame, 1.0


def expand_bbox(bbox, frame_shape, padding_ratio=FACE_DETECTION_PADDING_RATIO):
    frame_height, frame_width = frame_shape[:2]
    x1, y1, x2, y2 = bbox
    bbox_w = max(1, x2 - x1)
    bbox_h = max(1, y2 - y1)
    pad_x = int(bbox_w * padding_ratio)
    pad_y = int(bbox_h * padding_ratio)

    return (
        max(0, x1 - pad_x),
        max(0, y1 - pad_y),
        min(frame_width - 1, x2 + pad_x),
        min(frame_height - 1, y2 + pad_y),
    )


def remap_landmarks_to_frame(landmarks, crop_bbox, frame_shape):
    crop_x1, crop_y1, crop_x2, crop_y2 = crop_bbox
    crop_width = max(crop_x2 - crop_x1, 1)
    crop_height = max(crop_y2 - crop_y1, 1)
    frame_height, frame_width = frame_shape[:2]
    mapped = []
    frame_width = max(frame_width, 1)
    frame_height = max(frame_height, 1)

    for landmark in landmarks:
        mapped_x = (crop_x1 + landmark.x * crop_width) / frame_width
        mapped_y = (crop_y1 + landmark.y * crop_height) / frame_height
        mapped.append(
            SimpleNamespace(
                x=min(max(mapped_x, 0.0), 1.0),
                y=min(max(mapped_y, 0.0), 1.0),
                z=getattr(landmark, "z", 0.0),
            )
        )

    return mapped


def extract_face_landmarks_from_regions(rgb_frame, face_regions):
    frame_height, frame_width = rgb_frame.shape[:2]
    people_landmarks = []

    for face_region in face_regions:
        expanded_bbox = expand_bbox(face_region["bbox"], rgb_frame.shape)
        x1, y1, x2, y2 = expanded_bbox
        face_crop = rgb_frame[y1:y2, x1:x2]
        if face_crop.size == 0:
            continue

        crop_landmarks, processed_crop, crop_scale = detect_landmarks_direct_with_retry(face_crop)
        if not crop_landmarks:
            continue

        def landmark_area(landmarks):
            bbox = get_face_bbox(landmarks, processed_crop.shape[1], processed_crop.shape[0])
            return max(0, bbox[2] - bbox[0]) * max(0, bbox[3] - bbox[1])

        primary_landmarks = max(
            crop_landmarks,
            key=landmark_area,
        )
        mapped_landmarks = remap_landmarks_to_frame(
            primary_landmarks,
            expanded_bbox,
            rgb_frame.shape,
        )
        people_landmarks.append(
            {
                "landmarks": mapped_landmarks,
                "bbox": get_face_bbox(mapped_landmarks, frame_width, frame_height),
                "score": face_region.get("score", 0.0),
                "crop_scale": crop_scale,
            }
        )

    return people_landmarks


def get_face_landmarks_with_retry(rgb_frame):
    direct_regions = detect_face_regions(rgb_frame)
    if direct_regions:
        people_landmarks = extract_face_landmarks_from_regions(rgb_frame, direct_regions)
        if people_landmarks:
            return people_landmarks, rgb_frame, 1.0, len(direct_regions)

    retry_frame, retry_scale = upscale_frame_for_detection(rgb_frame)
    if retry_scale > 1.01:
        retry_regions = detect_face_regions(retry_frame)
        if retry_regions:
            people_landmarks = extract_face_landmarks_from_regions(retry_frame, retry_regions)
            if people_landmarks:
                # #region debug-point A:retry-upscale-hit
                report_debug_event(
                    "A",
                    "app.py:get_face_landmarks_with_retry:retry-upscale-hit",
                    "[DEBUG] face-first detection recovered after upscale retry",
                    {
                        "original_width": int(rgb_frame.shape[1]),
                        "original_height": int(rgb_frame.shape[0]),
                        "retry_width": int(retry_frame.shape[1]),
                        "retry_height": int(retry_frame.shape[0]),
                        "retry_scale": round(retry_scale, 3),
                        "detected_faces": len(retry_regions),
                        "landmark_faces": len(people_landmarks),
                    },
                )
                # #endregion
                return people_landmarks, retry_frame, retry_scale, len(retry_regions)

    fallback_landmarks, _, _ = detect_landmarks_direct_with_retry(rgb_frame)
    if fallback_landmarks:
        people_landmarks = [
            {
                "landmarks": landmarks,
                "bbox": get_face_bbox(landmarks, rgb_frame.shape[1], rgb_frame.shape[0]),
                "score": 0.0,
                "crop_scale": 1.0,
            }
            for landmarks in fallback_landmarks
        ]
        return people_landmarks, rgb_frame, 1.0, len(people_landmarks)

    return [], rgb_frame, 1.0, 0


def bbox_iou(box_a, box_b):
    if not box_a or not box_b:
        return 0.0

    ax1, ay1, ax2, ay2 = box_a
    bx1, by1, bx2, by2 = box_b
    inter_x1 = max(ax1, bx1)
    inter_y1 = max(ay1, by1)
    inter_x2 = min(ax2, bx2)
    inter_y2 = min(ay2, by2)
    inter_w = max(0, inter_x2 - inter_x1)
    inter_h = max(0, inter_y2 - inter_y1)
    inter_area = inter_w * inter_h
    area_a = max(0, ax2 - ax1) * max(0, ay2 - ay1)
    area_b = max(0, bx2 - bx1) * max(0, by2 - by1)
    denom = area_a + area_b - inter_area
    if denom <= 0:
        return 0.0
    return inter_area / denom


def assign_track_id(centroid, bbox, frame_shape, used_track_ids):
    global NEXT_TRACK_ID

    frame_height, frame_width = frame_shape[:2]
    max_distance = max(frame_width, frame_height) * TRACK_MATCH_DISTANCE_RATIO
    best_track_id = None
    best_iou = 0.0
    best_distance = max_distance

    for track_id, track_state in TRACKS.items():
        if track_id in used_track_ids:
            continue

        previous_centroid = track_state.get("centroid")
        if previous_centroid is None:
            continue

        distance = np.linalg.norm(np.array(centroid) - np.array(previous_centroid))
        iou = bbox_iou(bbox, track_state.get("bbox"))
        if iou >= TRACK_MIN_IOU and (iou > best_iou or (abs(iou - best_iou) < 1e-6 and distance < best_distance)):
            best_iou = iou
            best_distance = distance
            best_track_id = track_id
            continue

        if best_track_id is None and distance < best_distance:
            best_distance = distance
            best_track_id = track_id
        elif best_iou < TRACK_MIN_IOU and distance < best_distance:
            best_distance = distance
            best_track_id = track_id

    if best_track_id is None and len(TRACKS) < MAX_TRACKED_FACES:
        best_track_id = NEXT_TRACK_ID
        NEXT_TRACK_ID += 1
        TRACKS[best_track_id] = create_track_state()
    elif best_track_id is None:
        stale_track_id = max(TRACKS, key=lambda track_id: TRACKS[track_id]["misses"])
        best_track_id = stale_track_id
        TRACKS[best_track_id] = create_track_state()

    track_state = TRACKS[best_track_id]
    track_state["centroid"] = centroid
    track_state["bbox"] = bbox
    track_state["misses"] = 0
    track_state["last_seen"] = time.time()
    used_track_ids.add(best_track_id)

    return best_track_id, track_state


def cleanup_tracks():
    stale_track_ids = [
        track_id for track_id, track_state in TRACKS.items()
        if track_state["misses"] > TRACK_MAX_MISSES
    ]
    for track_id in stale_track_ids:
        TRACKS.pop(track_id, None)


def analyze_face(landmarks, frame_shape, track_id, track_state, use_trained=True):
    frame_height, frame_width = frame_shape[:2]
    left_ear = eye_aspect_ratio(landmarks, LEFT_EYE)
    right_ear = eye_aspect_ratio(landmarks, RIGHT_EYE)
    left_iris_ar = iris_aspect_ratio(landmarks, LEFT_IRIS)
    right_iris_ar = iris_aspect_ratio(landmarks, RIGHT_IRIS)
    ear = (left_ear + right_ear) / 2
    iris_ar = (left_iris_ar + right_iris_ar) / 2
    yaw, pitch, roll = head_pose(landmarks, frame_width, frame_height)

    track_state["ear_history"].append(ear)
    track_state["iris_history"].append(iris_ar)
    track_state["pose_history"].append((yaw, pitch, roll))
    ear_smooth = np.mean(track_state["ear_history"])
    iris_smooth = np.mean(track_state["iris_history"])
    pose_smooth = np.mean(track_state["pose_history"], axis=0)

    if track_state["calibration_mode"] or (
        len(track_state["calib_ear"]) < CALIBRATION_FRAMES and ear_smooth > 0.25
    ):
        track_state["calibration_mode"] = True
        track_state["calib_ear"].append(ear_smooth)
        track_state["calib_iris"].append(iris_smooth)
        if len(track_state["calib_ear"]) >= CALIBRATION_FRAMES:
            track_state["ear_thresh"] = np.mean(track_state["calib_ear"]) * 0.8
            track_state["iris_thresh"] = np.mean(track_state["calib_iris"]) * 0.8
            track_state["calibration_mode"] = False
            track_state["calib_ear"].clear()
            track_state["calib_iris"].clear()

    history_len = len(track_state["ear_history"])
    both_eyes_closed = (
        ear_smooth < track_state["ear_thresh"] and iris_smooth < track_state["iris_thresh"]
    )
    turned_away = abs(pose_smooth[1]) > 30
    if both_eyes_closed:
        track_state["closed_count"] += 1
        track_state["open_count"] = 0
    else:
        track_state["open_count"] += 1
        track_state["closed_count"] = 0

    if history_len <= 2:
        is_sleepy = (ear_smooth < STATIC_EAR_SLEEPY and iris_smooth < STATIC_IRIS_SLEEPY)
    else:
        is_sleepy = track_state["closed_count"] >= CONSEC_FRAMES_REQUIRED

    left_iris_x = np.mean([landmarks[i].x for i in LEFT_IRIS])
    right_iris_x = np.mean([landmarks[i].x for i in RIGHT_IRIS])
    left_eye_x_center = np.mean([landmarks[i].x for i in LEFT_EYE])
    right_eye_x_center = np.mean([landmarks[i].x for i in RIGHT_EYE])
    left_eye_w = max([landmarks[i].x for i in LEFT_EYE]) - min([landmarks[i].x for i in LEFT_EYE])
    right_eye_w = max([landmarks[i].x for i in RIGHT_EYE]) - min([landmarks[i].x for i in RIGHT_EYE])
    if left_eye_w < 1e-6:
        left_eye_w = 1e-6
    if right_eye_w < 1e-6:
        right_eye_w = 1e-6
    left_offset = (left_iris_x - left_eye_x_center) / left_eye_w
    right_offset = (right_iris_x - right_eye_x_center) / right_eye_w
    gaze_x = (left_offset + right_offset) / 2.0

    if use_trained and trained_clf is not None:
        try:
            features = np.array(
                [ear_smooth, iris_smooth, pose_smooth[0], pose_smooth[1], pose_smooth[2], gaze_x]
            )
            pred = trained_clf.predict([features])[0]
            pred_proba = trained_clf.predict_proba([features])[0]
            if pred == 1:
                status = "Focused"
                confidence = float(pred_proba[1])
            else:
                status = "Not Focused"
                confidence = float(pred_proba[0])
        except Exception as e:
            print(f"Model prediction error: {e}")
            if is_sleepy:
                status = "Not Focused (Sleepy)"
                confidence = 0.5
            elif turned_away:
                status = "Not Focused (Turned Away)"
                confidence = 0.6
            else:
                status = "Focused"
                confidence = 1.0
    else:
        if is_sleepy:
            status = "Not Focused (Sleepy)"
            confidence = 0.5
        elif turned_away:
            status = "Not Focused (Turned Away)"
            confidence = 0.6
        else:
            status = "Focused"
            confidence = 1.0

    bbox = get_face_bbox(landmarks, frame_width, frame_height)
    metrics = {
        "id": int(track_id),
        "label": f"Person {track_id}",
        "bbox": {
            "x1": int(bbox[0]),
            "y1": int(bbox[1]),
            "x2": int(bbox[2]),
            "y2": int(bbox[3]),
        },
        "ear": float(ear_smooth),
        "iris": float(iris_smooth),
        "yaw": float(pose_smooth[0]),
        "pitch": float(pose_smooth[1]),
        "roll": float(pose_smooth[2]),
        "gaze_x": float(gaze_x),
        "ear_thresh": float(track_state["ear_thresh"]),
        "iris_thresh": float(track_state["iris_thresh"]),
        "status": status,
        "confidence": confidence,
    }

    try:
        with open(LOG_FILE, "a", newline="") as f:
            writer = csv.writer(f)
            writer.writerow([
                time.time(),
                ear_smooth,
                iris_smooth,
                pose_smooth[0],
                pose_smooth[1],
                pose_smooth[2],
                f"P{track_id}:{status}",
                confidence,
            ])
    except Exception:
        pass

    return metrics


def draw_person_overlay(frame_bgr, person_metrics):
    bbox = person_metrics["bbox"]
    status = person_metrics["status"]
    color = (0, 255, 0) if status.startswith("Focused") else (0, 0, 255)

    x1, y1, x2, y2 = bbox["x1"], bbox["y1"], bbox["x2"], bbox["y2"]
    cv2.rectangle(frame_bgr, (x1, y1), (x2, y2), color, 2)

    label = (
        f"{person_metrics['label']} | {status} | "
        f"{person_metrics['confidence']:.2f}"
    )
    text_y = max(25, y1 - 10)
    cv2.putText(
        frame_bgr,
        label,
        (x1, text_y),
        cv2.FONT_HERSHEY_SIMPLEX,
        0.6,
        color,
        2,
    )

def detect_focus(frame, detailed=False, return_metrics=False, use_trained=True):
    # Normalize input from Gradio: accept None, file path, PIL image, or numpy array.
    from PIL import Image as PILImage

    def to_rgb(img):
        if img is None:
            return None
        if isinstance(img, str):
            arr = cv2.imread(img)
            if arr is None:
                return None
            return cv2.cvtColor(arr, cv2.COLOR_BGR2RGB)
        if isinstance(img, np.ndarray):
            if img.size == 0:
                return None
            # Gradio typically supplies RGB numpy arrays for images/webcam.
            return img
        if isinstance(img, PILImage.Image):
            return cv2.cvtColor(np.array(img), cv2.COLOR_RGB2RGB)
        return None

    rgb_frame = to_rgb(frame)
    if rgb_frame is None:
        # #region debug-point A:invalid-frame
        report_debug_event(
            "A",
            "app.py:detect_focus:invalid-frame",
            "[DEBUG] detect_focus received invalid frame",
            {
                "frame_type": type(frame).__name__ if frame is not None else "NoneType",
            },
        )
        # #endregion
        # return a small blank image with an error message
        blank = np.zeros((240, 320, 3), dtype=np.uint8)
        cv2.putText(blank, "No frame", (10, 120), cv2.FONT_HERSHEY_SIMPLEX, 1, (0,0,255), 2)
        return blank
    for track_state in TRACKS.values():
        track_state["misses"] += 1

    detected_faces, processed_rgb_frame, detection_scale, detector_face_count = get_face_landmarks_with_retry(rgb_frame)
    bgr = cv2.cvtColor(processed_rgb_frame, cv2.COLOR_RGB2BGR)
    frame_height, frame_width = bgr.shape[:2]

    # #region debug-point A:landmark-summary
    _face_scales = []
    for _entry in detected_faces[:5]:
        _bbox = _entry["bbox"]
        _bbox_w = max(0, _bbox[2] - _bbox[0])
        _bbox_h = max(0, _bbox[3] - _bbox[1])
        _face_scales.append(
            {
                "bbox_w": _bbox_w,
                "bbox_h": _bbox_h,
                "bbox_w_norm": round(_bbox_w / max(frame_width, 1), 4),
                "bbox_h_norm": round(_bbox_h / max(frame_height, 1), 4),
                "detector_score": round(float(_entry.get("score", 0.0)), 4),
                "crop_scale": round(float(_entry.get("crop_scale", 1.0)), 3),
            }
        )
    report_debug_event(
        "A",
        "app.py:detect_focus:landmark-summary",
        "[DEBUG] detect_focus landmark summary",
        {
            "frame_width": frame_width,
            "frame_height": frame_height,
            "detector_face_count": detector_face_count,
            "landmark_count": len(detected_faces),
            "face_scales": _face_scales,
            "use_tasks": USE_TASKS,
            "detection_scale": round(detection_scale, 3),
        },
    )
    # #endregion

    if not detected_faces:
        cleanup_tracks()
        # #region debug-point B:no-face
        report_debug_event(
            "B",
            "app.py:detect_focus:no-face",
            "[DEBUG] detect_focus no face detected",
            {
                "frame_width": frame_width,
                "frame_height": frame_height,
                "detector_face_count": detector_face_count,
                "tracked_faces": len(TRACKS),
            },
        )
        # #endregion
        font_scale = max(0.6, min(3.0, frame_height / 480.0))
        thickness = max(1, int(round(frame_height / 240.0)))
        cv2.putText(
            bgr,
            "No face detected",
            (30, int(30 * (frame_height / 480.0))),
            cv2.FONT_HERSHEY_SIMPLEX,
            font_scale,
            (0, 0, 255),
            thickness,
        )
        if return_metrics:
            return bgr, {"status": "no_face", "people": []}
        return bgr

    used_track_ids = set()
    people_metrics = []
    sorted_detected_faces = sorted(
        detected_faces,
        key=lambda entry: entry["bbox"][0],
    )

    for detected_face in sorted_detected_faces:
        landmarks = detected_face["landmarks"]
        bbox = detected_face["bbox"]
        centroid = ((bbox[0] + bbox[2]) / 2.0, (bbox[1] + bbox[3]) / 2.0)
        track_id, track_state = assign_track_id(centroid, bbox, bgr.shape, used_track_ids)
        person_metrics = analyze_face(
            landmarks,
            bgr.shape,
            track_id,
            track_state,
            use_trained=use_trained,
        )
        people_metrics.append(person_metrics)
        draw_person_overlay(bgr, person_metrics)

    cleanup_tracks()

    focused_count = sum(1 for person in people_metrics if person["status"].startswith("Focused"))
    summary = {
        "status": "multi_face",
        "people_count": len(people_metrics),
        "focused_count": focused_count,
        "not_focused_count": len(people_metrics) - focused_count,
        "people": people_metrics,
    }

    if return_metrics:
        return bgr, summary
    return bgr

def run_inference(frame, use_trained):
    result = detect_focus(frame, return_metrics=True, use_trained=use_trained)
    if isinstance(result, tuple):
        output_frame, metrics = result
        people = metrics.get("people", [])
        if not people:
            status_line = "Tidak ada wajah terdeteksi."
        else:
            status_line = (
                f"Terdeteksi {metrics.get('people_count', len(people))} orang | "
                f"Focused: {metrics.get('focused_count', 0)} | "
                f"Not Focused: {metrics.get('not_focused_count', 0)}"
            )
        return output_frame, status_line, format_metrics(metrics), build_people_rows(metrics)
    return result, "Belum ada hasil.", "{}", []


def run_webcam_inference(frame, use_trained):
    optimized_frame = optimize_frame_for_webcam(frame)
    result = detect_focus(optimized_frame, return_metrics=True, use_trained=use_trained)
    if isinstance(result, tuple):
        output_frame, metrics = result
        append_record_events(metrics)
        people = metrics.get("people", [])
        if not people:
            status_line = "Tidak ada wajah terdeteksi."
        else:
            status_line = (
                f"Terdeteksi {metrics.get('people_count', len(people))} orang | "
                f"Focused: {metrics.get('focused_count', 0)} | "
                f"Not Focused: {metrics.get('not_focused_count', 0)}"
            )
        return (
            output_frame,
            status_line,
            format_metrics(metrics),
            build_people_rows(metrics),
            build_recording_status(),
            build_record_event_rows(),
            build_record_summary_rows(),
        )
    return result, "Belum ada hasil.", "{}", [], build_recording_status(), build_record_event_rows(), build_record_summary_rows()


def bind_webcam_inference(webcam_component, use_trained_component, outputs):
    if hasattr(webcam_component, "stream"):
        try:
            webcam_component.stream(
                fn=run_webcam_inference,
                inputs=[webcam_component, use_trained_component],
                outputs=outputs,
                trigger_mode="always_last",
                concurrency_limit=1,
                queue=False,
                time_limit=None,
                stream_every=WEBCAM_STREAM_EVERY,
            )
        except TypeError:
            webcam_component.stream(
                fn=run_webcam_inference,
                inputs=[webcam_component, use_trained_component],
                outputs=outputs,
            )
    webcam_component.change(
        fn=run_webcam_inference,
        inputs=[webcam_component, use_trained_component],
        outputs=outputs,
    )


# FastAPI Endpoints
@app.get("/health")
async def health():
    return {"status": "ok", "model_loaded": trained_clf is not None}


@app.post("/focus/analyze-frame")
async def analyze_frame(request: FrameRequest):
    try:
        # Decode base64 image
        img_data = base64.b64decode(request.image_base64.split(",")[-1])
        nparr = np.frombuffer(img_data, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        if img is None:
            raise HTTPException(status_code=400, detail="Invalid image")

        # #region debug-point C:incoming-frame
        report_debug_event(
            "C",
            "app.py:analyze_frame:incoming-frame",
            "[DEBUG] analyze_frame incoming frame",
            {
                "encoded_bytes": len(img_data),
                "decoded_width": int(img.shape[1]),
                "decoded_height": int(img.shape[0]),
                "use_trained": bool(request.use_trained),
            },
        )
        # #endregion
        
        # Convert to RGB
        img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        
        # Analyze
        result_frame, metrics = detect_focus(img_rgb, return_metrics=True, use_trained=request.use_trained)
        append_record_events(metrics)

        # #region debug-point D:analyze-result
        report_debug_event(
            "D",
            "app.py:analyze_frame:analyze-result",
            "[DEBUG] analyze_frame result summary",
            {
                "status": metrics.get("status") if isinstance(metrics, dict) else "unknown",
                "people_count": metrics.get("people_count", 0) if isinstance(metrics, dict) else 0,
                "focused_count": metrics.get("focused_count", 0) if isinstance(metrics, dict) else 0,
                "not_focused_count": metrics.get("not_focused_count", 0) if isinstance(metrics, dict) else 0,
            },
        )
        # #endregion
        
        # Encode result frame to base64
        _, buffer = cv2.imencode('.jpg', result_frame)
        result_base64 = base64.b64encode(buffer).decode('utf-8')
        
        return {
            "success": True,
            "metrics": metrics,
            "annotated_image_base64": f"data:image/jpeg;base64,{result_base64}"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/focus/record/start")
async def start_record():
    status, events, summary = start_recording()
    return {
        "success": True,
        "status": status,
        "events": events,
        "summary": summary
    }


@app.post("/focus/record/stop")
async def stop_record():
    status, events, summary = stop_recording()
    return {
        "success": True,
        "status": status,
        "events": events,
        "summary": summary,
        "full_events": RECORD_EVENTS,
        "full_summary": build_record_summary_rows()
    }


@app.post("/focus/record/clear")
async def clear_record():
    status, events, summary = clear_recording()
    return {
        "success": True,
        "status": status,
        "events": events,
        "summary": summary
    }


@app.get("/focus/record/status")
async def get_record_status():
    return {
        "success": True,
        "status": build_recording_status(),
        "events": build_record_event_rows(),
        "summary": build_record_summary_rows(),
        "is_recording": RECORDING_ACTIVE
    }


@app.get("/focus/record/export")
async def export_record():
    try:
        message, export_path = export_record_to_excel()
        if not os.path.exists(export_path):
            raise HTTPException(status_code=404, detail="Export file not found")
        
        def iterfile():
            with open(export_path, mode="rb") as file_like:
                yield from file_like
        
        filename = os.path.basename(export_path)
        return StreamingResponse(
            iterfile(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/focus/reset")
async def reset_focus():
    reset_tracking_state()
    clear_recording_state()
    return {"success": True, "message": "State reset successfully"}


# Gradio UI
use_blocks = hasattr(gr, "Blocks")

if use_blocks:
    with gr.Blocks(title="Focus Detector") as demo:
        gr.Markdown("# Focus Detector Webcam Test")
        gr.Markdown(
            "UI sederhana untuk mencoba model AI focus detector dengan webcam atau gambar statis."
        )
        backend_name = "MediaPipe Tasks" if USE_TASKS else "MediaPipe FaceMesh"
        model_name = "focus_model.pkl" if trained_clf is not None else "heuristic fallback"
        gr.Markdown(
            f"Backend: **{backend_name}**  \n"
            f"Mode prediksi default: **{model_name}**"
        )

        with gr.Tabs():
            with gr.TabItem("Webcam"):
                gr.Markdown(
                    "Izinkan akses kamera di browser, lalu lihat hasil prediksi pada panel kanan."
                )

                with gr.Row():
                    use_trained_webcam = gr.Checkbox(
                        label="Gunakan model terlatih",
                        value=trained_clf is not None,
                        interactive=trained_clf is not None,
                    )
                    reset_webcam = gr.Button("Reset state")
                    start_record_btn = gr.Button("Start Record")
                    stop_record_btn = gr.Button("Stop Record")
                    clear_record_btn = gr.Button("Clear Record")
                    export_excel_btn = gr.Button("Export Excel")

                with gr.Row():
                    with gr.Column():
                        cam = gr.Image(
                            sources=["webcam"],
                            label="Webcam Input",
                            type="numpy",
                            streaming=True,
                            height=360,
                        )
                    with gr.Column():
                        out_cam = gr.Image(label="Detection Output", height=360)
                        webcam_status = gr.Textbox(label="Ringkasan", interactive=False)
                        record_status = gr.Textbox(
                            label="Status Record",
                            value=build_recording_status(),
                            interactive=False,
                        )
                        webcam_table = gr.Dataframe(
                            headers=["ID", "Label", "Status", "Confidence", "BBox", "Gaze X"],
                            datatype=["number", "str", "str", "number", "str", "number"],
                            row_count=5,
                            col_count=(6, "fixed"),
                            interactive=False,
                            label="Tabel Deteksi Orang",
                        )
                        webcam_metrics = gr.Textbox(label="Metrics", lines=12, interactive=False)

                with gr.Row():
                    record_events_table = gr.Dataframe(
                        headers=["Timestamp", "ID", "Label", "Status", "Confidence"],
                        datatype=["str", "number", "str", "str", "number"],
                        row_count=10,
                        col_count=(5, "fixed"),
                        interactive=False,
                        label="Log Record per Timestamp",
                        value=build_record_event_rows(),
                    )

                with gr.Row():
                    record_summary_table = gr.Dataframe(
                        headers=["ID", "Label", "Focused", "Not Focused", "Total", "First Seen", "Last Seen"],
                        datatype=["number", "str", "number", "number", "number", "str", "str"],
                        row_count=5,
                        col_count=(7, "fixed"),
                        interactive=False,
                        label="Rekap Record per ID",
                        value=build_record_summary_rows(),
                    )

                with gr.Row():
                    export_status = gr.Textbox(label="Status Export", interactive=False)
                    export_file = gr.File(label="File Excel", interactive=False)

                bind_webcam_inference(
                    cam,
                    use_trained_webcam,
                    [
                        out_cam,
                        webcam_status,
                        webcam_metrics,
                        webcam_table,
                        record_status,
                        record_events_table,
                        record_summary_table,
                    ],
                )
                reset_webcam.click(
                    fn=reset_webcam_state,
                    outputs=[
                        webcam_status,
                        webcam_metrics,
                        webcam_table,
                        record_status,
                        record_events_table,
                        record_summary_table,
                    ],
                    queue=False,
                )
                start_record_btn.click(
                    fn=start_recording,
                    outputs=[record_status, record_events_table, record_summary_table],
                    queue=False,
                )
                stop_record_btn.click(
                    fn=stop_recording,
                    outputs=[record_status, record_events_table, record_summary_table],
                    queue=False,
                )
                clear_record_btn.click(
                    fn=clear_recording,
                    outputs=[record_status, record_events_table, record_summary_table],
                    queue=False,
                )
                export_excel_btn.click(
                    fn=export_record_to_excel,
                    outputs=[export_status, export_file],
                    queue=False,
                )

            with gr.TabItem("Upload Image"):
                gr.Markdown("Unggah gambar untuk cek hasil model tanpa webcam.")

                with gr.Row():
                    use_trained_upload = gr.Checkbox(
                        label="Gunakan model terlatih",
                        value=trained_clf is not None,
                        interactive=trained_clf is not None,
                    )
                    reset_upload = gr.Button("Reset state")

                with gr.Row():
                    with gr.Column():
                        upload = gr.Image(type="numpy", label="Upload Image")
                    with gr.Column():
                        out_up = gr.Image(label="Detection Output")
                        upload_status = gr.Textbox(label="Ringkasan", interactive=False)
                        upload_table = gr.Dataframe(
                            headers=["ID", "Label", "Status", "Confidence", "BBox", "Gaze X"],
                            datatype=["number", "str", "str", "number", "str", "number"],
                            row_count=5,
                            col_count=(6, "fixed"),
                            interactive=False,
                            label="Tabel Deteksi Orang",
                        )

                with gr.Row():
                    out_metrics = gr.Textbox(label="Metrics", lines=12, interactive=False)

                upload.change(
                    fn=run_inference,
                    inputs=[upload, use_trained_upload],
                    outputs=[out_up, upload_status, out_metrics, upload_table],
                )
                reset_upload.click(
                    fn=reset_runtime_state,
                    outputs=[upload_status, out_metrics, upload_table],
                )


# Mount Gradio to FastAPI
from fastapi.middleware.wsgi import WSGIMiddleware
gradio_app = gr.mount_gradio_app(app, demo, path="/gradio")


if __name__ == "__main__":
    import uvicorn
    server_name = os.getenv("GRADIO_SERVER_NAME", "0.0.0.0")
    server_port = int(os.getenv("GRADIO_SERVER_PORT", "7861"))
    uvicorn.run(app, host=server_name, port=server_port)
